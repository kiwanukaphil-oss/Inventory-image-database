"""
build_review.py — Generate a self-contained image-review gallery (review.html).

Reads pants_products.xlsx (the "Image map" sheet as the per-image review grain,
plus "Variants" for read-only price/stock context), resolves every image_file to
its actual folder by scanning the category directories, and emits a single
double-clickable review.html with embedded data. No server, works offline.

The gallery supports inline edit-back: edits persist in the browser (localStorage)
and the Export button downloads a corrections CSV in the SAME schema as
corrections_for_approval.csv, so it plugs straight into the existing apply flow.
"""

import json
import os
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "pants_products.xlsx")
IMAGE_DIRS = ["Cargo Pants", "Flat_front", "Formal_pants", "Khaki", "Linen", "Relaxed_Draw string"]
EDITABLE_FIELDS = ["brand", "color", "size", "style", "material"]


def build_filename_index():
    """Map each image filename -> its relative 'Folder/filename' path.

    image_file in the spreadsheet is a bare filename, so we scan every category
    folder to find where each image actually lives. Returns the index plus the
    set of files seen, so the caller can detect orphans (on disk, not in sheet).
    """
    index = {}
    for d in IMAGE_DIRS:
        full = os.path.join(ROOT, d)
        if not os.path.isdir(full):
            continue
        for fn in os.listdir(full):
            if fn.lower().endswith((".jpg", ".jpeg", ".png", ".webp")):
                # Prefer first occurrence; warn on duplicate basenames across folders.
                index.setdefault(fn, f"{d}/{fn}")
    return index


def load_variant_context():
    """Build sku -> {price, cost_price, stock_quantity, reorder_level} for read-only display."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    rows = list(wb["Variants"].iter_rows(values_only=True))
    hdr = list(rows[0])
    ctx = {}
    for r in rows[1:]:
        rec = dict(zip(hdr, r))
        sku = rec.get("sku")
        if sku:
            ctx[sku] = {
                "price": rec.get("price"),
                "cost_price": rec.get("cost_price"),
                "stock_quantity": rec.get("stock_quantity"),
                "reorder_level": rec.get("reorder_level"),
            }
    return ctx


def load_image_map(file_index, variant_ctx):
    """Join Image map rows to resolved image paths + variant context.

    Each emitted record carries the original field values (used as the 'before'
    baseline for change export) and a resolved relative image path (or null when
    the image is missing on disk, which the UI flags).
    """
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    rows = list(wb["Image map"].iter_rows(values_only=True))
    hdr = list(rows[0])
    records = []
    used = set()
    for i, r in enumerate(rows[1:]):
        rec = dict(zip(hdr, r))
        fn = rec.get("image_file")
        rel = file_index.get(fn) if fn else None
        if rel:
            used.add(fn)
        sku = rec.get("variant_sku")
        records.append({
            "id": i,
            "product_name": rec.get("product_name"),
            "variant_sku": sku,
            "size": _s(rec.get("size")),
            "color": _s(rec.get("color")),
            "brand": _s(rec.get("brand")),
            "style": _s(rec.get("style")),
            "material": _s(rec.get("material")),
            "image_file": fn,
            "rel_path": rel,            # null => missing on disk
            "ctx": variant_ctx.get(sku, {}),
        })
    return records, used


def _s(v):
    """Normalise a cell to a trimmed string ('' for None) so the UI never shows 'None'."""
    return "" if v is None else str(v).strip()


def main():
    file_index = build_filename_index()
    variant_ctx = load_variant_context()
    records, used = load_image_map(file_index, variant_ctx)

    missing = [r for r in records if not r["rel_path"]]
    orphans = sorted(set(file_index) - used)  # on disk but not referenced by any row

    payload = {
        "records": records,
        "editableFields": EDITABLE_FIELDS,
        "orphans": [file_index[o] for o in orphans],
        "stats": {
            "total": len(records),
            "missing": len(missing),
            "orphans": len(orphans),
        },
    }

    html = HTML_TEMPLATE.replace("/*__DATA__*/", json.dumps(payload, ensure_ascii=False))
    out = os.path.join(ROOT, "review.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Wrote {out}")
    print(f"  {payload['stats']['total']} rows | "
          f"{payload['stats']['missing']} missing images | "
          f"{payload['stats']['orphans']} orphan images (on disk, not in sheet)")
    if missing:
        print("  Missing image files (first 10):")
        for r in missing[:10]:
            print("    -", r["image_file"])


# The HTML is kept as one template string; /*__DATA__*/ is replaced with JSON above.
HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Inventory Image Review</title>
<style>
  :root { --bg:#0f1115; --panel:#171a21; --line:#2a2f3a; --txt:#e6e9ef; --muted:#9aa3b2;
          --accent:#4f8cff; --edit:#3a2f00; --flag:#3a0f12; --ok:#13351c; }
  * { box-sizing:border-box; }
  body { margin:0; background:var(--bg); color:var(--txt);
         font:14px/1.4 system-ui,Segoe UI,Roboto,sans-serif; }
  header { position:sticky; top:0; z-index:20; background:var(--panel);
           border-bottom:1px solid var(--line); padding:10px 14px; }
  h1 { font-size:16px; margin:0 0 8px; }
  .controls { display:flex; flex-wrap:wrap; gap:8px; align-items:center; }
  input, select, button { background:#0d0f14; color:var(--txt); border:1px solid var(--line);
           border-radius:6px; padding:6px 8px; font:inherit; }
  input[type=text]{ min-width:200px; }
  button { cursor:pointer; }
  button.primary { background:var(--accent); border-color:var(--accent); color:#fff; font-weight:600; }
  button.ghost:hover { border-color:var(--accent); }
  .stats { color:var(--muted); margin-left:auto; font-size:12px; text-align:right; }
  .grid { padding:12px; display:grid; grid-template-columns:repeat(auto-fill,minmax(300px,1fr)); gap:12px; }
  .card { background:var(--panel); border:1px solid var(--line); border-radius:10px; overflow:hidden;
          display:flex; flex-direction:column; }
  .card.edited { border-color:#a98b00; box-shadow:0 0 0 1px #a98b00 inset; }
  .card.flagged { border-color:#d04b53; box-shadow:0 0 0 1px #d04b53 inset; }
  .card.missing .thumb { background:repeating-linear-gradient(45deg,#2a0f12,#2a0f12 10px,#1c0a0c 10px,#1c0a0c 20px); }
  .thumb { aspect-ratio:1/1; background:#000; display:flex; align-items:center; justify-content:center;
           cursor:zoom-in; overflow:hidden; position:relative; }
  .thumb img { width:100%; height:100%; object-fit:contain; }
  .thumb .nomedia { color:var(--muted); font-size:12px; text-align:center; padding:8px; }
  .badge { position:absolute; top:6px; left:6px; background:#000a; border:1px solid var(--line);
           border-radius:6px; padding:2px 6px; font-size:11px; }
  .body { padding:8px 10px; display:flex; flex-direction:column; gap:6px; }
  .skuline { font-size:11px; color:var(--muted); word-break:break-all; }
  .row { display:grid; grid-template-columns:64px 1fr; gap:6px; align-items:center; }
  .row label { color:var(--muted); font-size:12px; }
  .row input.field { padding:4px 6px; width:100%; }
  .row input.field.changed { background:var(--edit); border-color:#a98b00; }
  .ctx { font-size:11px; color:var(--muted); display:flex; gap:10px; flex-wrap:wrap; }
  .cardfoot { display:flex; gap:6px; align-items:center; padding:6px 10px; border-top:1px solid var(--line); }
  .cardfoot input.note { flex:1; padding:4px 6px; font-size:12px; }
  .flagbtn { font-size:12px; }
  .flagbtn.on { background:var(--flag); border-color:#d04b53; color:#ffb3b8; }
  /* lightbox */
  #lb { position:fixed; inset:0; background:#000d; z-index:50; display:none;
        align-items:center; justify-content:center; flex-direction:column; }
  #lb.open { display:flex; }
  #lb img { max-width:92vw; max-height:82vh; object-fit:contain; }
  #lb .cap { color:#ddd; margin-top:10px; font-size:13px; text-align:center; max-width:90vw; }
  #lb .hint { color:#888; font-size:12px; margin-top:4px; }
  .pill { font-size:11px; padding:1px 6px; border-radius:10px; border:1px solid var(--line); color:var(--muted); }
</style>
</head>
<body>
<header>
  <h1>Inventory Image Review</h1>
  <div class="controls">
    <input id="q" type="text" placeholder="Search brand / color / sku / file…" autofocus>
    <select id="cat"><option value="">All categories</option></select>
    <label class="pill"><input id="onlyEdited" type="checkbox"> edited</label>
    <label class="pill"><input id="onlyFlagged" type="checkbox"> flagged</label>
    <label class="pill"><input id="onlyMissing" type="checkbox"> missing img</label>
    <button class="ghost" id="bigger">A+</button>
    <button class="ghost" id="smaller">A−</button>
    <button class="primary" id="export">Export corrections CSV</button>
    <button class="ghost" id="reset">Reset edits</button>
    <span class="stats" id="stats"></span>
  </div>
</header>

<div class="grid" id="grid"></div>

<div id="lb">
  <img id="lbimg" alt="">
  <div class="cap" id="lbcap"></div>
  <div class="hint">← / → to move · Esc to close</div>
</div>

<script>
const PAYLOAD = /*__DATA__*/;
const LS_KEY = "inv_review_edits_v1";   // { [id]: { fields:{field:newval}, flag:bool, note:str } }
const EF = PAYLOAD.editableFields;

let edits = loadEdits();
let cardSize = 300;
let visibleOrder = [];   // record ids currently shown, for lightbox navigation

function loadEdits(){ try { return JSON.parse(localStorage.getItem(LS_KEY)) || {}; } catch(e){ return {}; } }
function saveEdits(){ localStorage.setItem(LS_KEY, JSON.stringify(edits)); }
function editOf(id){ return edits[id] || (edits[id] = { fields:{}, flag:false, note:"" }); }
function curVal(rec, f){ const e = edits[rec.id]; return (e && f in e.fields) ? e.fields[f] : rec[f]; }
function isChanged(rec, f){ const e = edits[rec.id]; return !!(e && f in e.fields && e.fields[f] !== (rec[f]||"")); }
function rowEdited(rec){ const e=edits[rec.id]; return !!(e && Object.keys(e.fields).some(f=>e.fields[f]!==(rec[f]||""))); }
function rowFlagged(rec){ const e=edits[rec.id]; return !!(e && e.flag); }

// Populate category filter from the data.
const cats = [...new Set(PAYLOAD.records.map(r=>r.product_name).filter(Boolean))].sort();
const catSel = document.getElementById("cat");
cats.forEach(c=>{ const o=document.createElement("option"); o.value=c; o.textContent=c; catSel.appendChild(o); });

const grid = document.getElementById("grid");

function matches(rec){
  const q = document.getElementById("q").value.trim().toLowerCase();
  const cat = catSel.value;
  if (cat && rec.product_name !== cat) return false;
  if (document.getElementById("onlyEdited").checked && !rowEdited(rec)) return false;
  if (document.getElementById("onlyFlagged").checked && !rowFlagged(rec)) return false;
  if (document.getElementById("onlyMissing").checked && rec.rel_path) return false;
  if (!q) return true;
  const hay = [rec.brand,rec.color,rec.size,rec.style,rec.material,rec.variant_sku,rec.image_file,rec.product_name]
              .join(" ").toLowerCase();
  return hay.includes(q);
}

function render(){
  const shown = PAYLOAD.records.filter(matches);
  visibleOrder = shown.map(r=>r.id);
  grid.style.gridTemplateColumns = `repeat(auto-fill,minmax(${cardSize}px,1fr))`;
  grid.innerHTML = "";
  for (const rec of shown) grid.appendChild(card(rec));
  updateStats(shown.length);
}

function card(rec){
  const el = document.createElement("div");
  el.className = "card" + (rowEdited(rec)?" edited":"") + (rowFlagged(rec)?" flagged":"") + (rec.rel_path?"":" missing");
  el.dataset.id = rec.id;

  // thumbnail (lazy) or missing placeholder
  const thumb = document.createElement("div");
  thumb.className = "thumb";
  if (rec.rel_path){
    const img = document.createElement("img");
    img.loading = "lazy";
    img.src = encodeURI(rec.rel_path);
    img.alt = rec.image_file;
    thumb.appendChild(img);
    thumb.onclick = ()=>openLightbox(rec.id);
    const b = document.createElement("div"); b.className="badge"; b.textContent = rec.size||"?"; thumb.appendChild(b);
  } else {
    const n = document.createElement("div"); n.className="nomedia";
    n.textContent = "image not found:\n"+(rec.image_file||"(blank)"); thumb.appendChild(n);
  }
  el.appendChild(thumb);

  const body = document.createElement("div"); body.className="body";
  const sku = document.createElement("div"); sku.className="skuline";
  sku.textContent = (rec.product_name||"") + " · " + (rec.variant_sku||"(no sku)");
  body.appendChild(sku);

  for (const f of EF){
    const row = document.createElement("div"); row.className="row";
    const lab = document.createElement("label"); lab.textContent = f; row.appendChild(lab);
    const inp = document.createElement("input");
    inp.className = "field" + (isChanged(rec,f)?" changed":"");
    inp.value = curVal(rec, f) || "";
    inp.oninput = ()=>{
      const e = editOf(rec.id);
      if (inp.value === (rec[f]||"")) delete e.fields[f]; else e.fields[f] = inp.value;
      inp.classList.toggle("changed", isChanged(rec,f));
      el.classList.toggle("edited", rowEdited(rec));
      saveEdits(); updateStats();
    };
    row.appendChild(inp); body.appendChild(row);
  }

  // read-only variant context
  const ctx = document.createElement("div"); ctx.className="ctx";
  const c = rec.ctx||{};
  ctx.innerHTML = `<span>price: ${fmt(c.price)}</span><span>cost: ${fmt(c.cost_price)}</span>`
                + `<span>stock: ${fmt(c.stock_quantity)}</span><span>reorder: ${fmt(c.reorder_level)}</span>`;
  body.appendChild(ctx);
  el.appendChild(body);

  const foot = document.createElement("div"); foot.className="cardfoot";
  const flag = document.createElement("button");
  flag.className = "flagbtn ghost" + (rowFlagged(rec)?" on":"");
  flag.textContent = rowFlagged(rec) ? "⚑ flagged" : "⚐ flag";
  flag.onclick = ()=>{ const e=editOf(rec.id); e.flag=!e.flag; saveEdits();
                       flag.classList.toggle("on",e.flag); flag.textContent=e.flag?"⚑ flagged":"⚐ flag";
                       el.classList.toggle("flagged",e.flag); updateStats(); };
  foot.appendChild(flag);
  const note = document.createElement("input"); note.className="note"; note.placeholder="note / evidence…";
  note.value = (edits[rec.id]&&edits[rec.id].note)||"";
  note.oninput = ()=>{ editOf(rec.id).note = note.value; saveEdits(); };
  foot.appendChild(note);
  el.appendChild(foot);
  return el;
}

function fmt(v){ return (v===null||v===undefined||v==="") ? "—" : v; }

function updateStats(shownCount){
  const edited = PAYLOAD.records.filter(rowEdited).length;
  const flagged = PAYLOAD.records.filter(rowFlagged).length;
  const s = PAYLOAD.stats;
  const shown = (shownCount===undefined) ? document.querySelectorAll(".card").length : shownCount;
  document.getElementById("stats").innerHTML =
    `${shown} shown / ${s.total} rows · <b>${edited}</b> edited · <b>${flagged}</b> flagged · `
    + `${s.missing} missing img · ${s.orphans} orphan img`;
}

/* ---------- lightbox ---------- */
const lb = document.getElementById("lb"), lbimg = document.getElementById("lbimg"), lbcap = document.getElementById("lbcap");
let lbId = null;
function openLightbox(id){
  const rec = PAYLOAD.records.find(r=>r.id===id); if(!rec||!rec.rel_path) return;
  lbId = id; lbimg.src = encodeURI(rec.rel_path);
  lbcap.textContent = `${rec.image_file}  —  ${rec.brand} / ${rec.color} / ${rec.size}`;
  lb.classList.add("open");
}
function moveLightbox(d){
  const i = visibleOrder.indexOf(lbId); if(i<0) return;
  let j=i+d; while(j>=0 && j<visibleOrder.length){
    const rec = PAYLOAD.records.find(r=>r.id===visibleOrder[j]);
    if(rec && rec.rel_path){ openLightbox(rec.id); return; } j+=d;
  }
}
lb.onclick = e=>{ if(e.target===lb) lb.classList.remove("open"); };
document.addEventListener("keydown", e=>{
  if(!lb.classList.contains("open")) return;
  if(e.key==="Escape") lb.classList.remove("open");
  else if(e.key==="ArrowRight") moveLightbox(1);
  else if(e.key==="ArrowLeft") moveLightbox(-1);
});

/* ---------- CSV export (matches corrections_for_approval.csv schema) ---------- */
function csvEscape(v){ v=(v===null||v===undefined)?"":String(v);
  return /[",\n]/.test(v) ? '"'+v.replace(/"/g,'""')+'"' : v; }
function exportCSV(){
  const cols = ["decision","confidence","category","image_file","field","current_value","proposed_value","your_correct_value","evidence"];
  const lines = [cols.join(",")];
  let n=0;
  for (const rec of PAYLOAD.records){
    const e = edits[rec.id]; if(!e) continue;
    const imgRef = rec.rel_path || rec.image_file || "";
    // one row per changed field
    for (const f of Object.keys(e.fields)){
      if (e.fields[f] === (rec[f]||"")) continue;
      lines.push([ "EDIT","User", rec.product_name, imgRef, f, rec[f]||"", "", e.fields[f], e.note||"" ].map(csvEscape).join(","));
      n++;
    }
    // flagged with no field change -> a FLAG row so it isn't lost
    const hasFieldChange = Object.keys(e.fields).some(f=>e.fields[f]!==(rec[f]||""));
    if (e.flag && !hasFieldChange){
      lines.push([ "FLAG","User", rec.product_name, imgRef, "", "", "", "", e.note||"needs review" ].map(csvEscape).join(","));
      n++;
    }
  }
  if(!n){ alert("No edits or flags to export yet."); return; }
  const blob = new Blob([lines.join("\n")], {type:"text/csv;charset=utf-8"});
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = "review_corrections.csv";
  a.click();
}

/* ---------- wiring ---------- */
["q","cat","onlyEdited","onlyFlagged","onlyMissing"].forEach(id=>{
  document.getElementById(id).addEventListener("input", render);
});
document.getElementById("export").onclick = exportCSV;
document.getElementById("reset").onclick = ()=>{
  if(confirm("Discard ALL in-browser edits and flags? (does not touch your files)")){
    edits={}; saveEdits(); render();
  }
};
document.getElementById("bigger").onclick = ()=>{ cardSize=Math.min(560,cardSize+40); render(); };
document.getElementById("smaller").onclick = ()=>{ cardSize=Math.max(200,cardSize-40); render(); };

render();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
